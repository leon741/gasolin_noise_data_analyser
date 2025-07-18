import sys
from pathlib import Path

from PyQt6.QtWidgets import (
    QApplication, QWidget, QPushButton, QLabel,
    QVBoxLayout, QHBoxLayout, QProgressBar, QLineEdit,
    QComboBox
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from openpyxl.styles import Alignment

import pandas as pd
import matplotlib.pyplot as plt
import os
from pathlib import Path 
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side
from matplotlib.ticker import MaxNLocator
import bisect

plt.rcParams['legend.fontsize'] = 5

def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

def plot_acceleration_graph(ax, time_list, acc_list, start_point):
    if not (0 <= start_point < len(time_list)):
        raise IndexError(f"start_point {start_point} over time_list 长度")
    start_idx = max(0, start_point - 24)
    x = time_list[start_idx:]
    y = acc_list[start_idx:]

    ax.plot(x, y, label='Accelerations', color='blue', linewidth=0.5)
    ax.axvline(x=time_list[start_point], color='red', linestyle='--',
               label='Peak Point', linewidth=1)
    zero_point = x[0]
    x_labels = [str(round(i - zero_point, 3)) for i in x]
    ax.set_xticks(x)
    ax.set_xticklabels(x_labels, fontsize=5, rotation=0)
    ax.set_xlabel("Time (s)", fontsize=6)
    ax.set_ylabel("ACC_X (G)", fontsize=6)
    ax.set_title("Acceleration Graph", fontsize=6)
    ax.tick_params(axis='both', labelsize=5)
    ax.grid(True)
    ax.legend(fontsize=5)
    ax.xaxis.set_major_locator(MaxNLocator(integer=True))

def plot_rear_center_noise(ax, time_list, noise_list, start_point, before_one_min_val, after_one_min_val, terminate_sec):
    start_idx = max(0, start_point - 24)
    t0 = time_list[start_point]  # 设为0秒的参考点（peak）

    # 计算终止点索引
    if terminate_sec is None:
        end_idx = len(time_list)
    else:
        terminate = t0 + float(terminate_sec)
        end_idx = bisect.bisect_left(time_list, terminate)

    # 数据截取
    plot_time = time_list[start_idx:end_idx]
    plot_noise = noise_list[start_idx:end_idx]
    shifted_time = [t - t0 for t in plot_time]
    shifted_all_time = [t - t0 for t in time_list]

    # 绘图
    ax.plot(shifted_time, plot_noise, label='Rear Seat Center Noise (dB)', color='blue', linewidth=0.5)
    ax.axvline(x=0, color='red', linestyle='--', label='Peak Point', linewidth=1)
    ax.hlines(y=before_one_min_val, xmin=0, xmax=1, colors='purple', linestyle='-', linewidth=1)
    ax.hlines(y=after_one_min_val, xmin=1, xmax=shifted_time[-1], colors='purple', linestyle='-', linewidth=0.5)
    ax.vlines(x=1, ymin=after_one_min_val, ymax=before_one_min_val, colors='purple', linestyle='-', label='Reference Line', linewidth=1)

    # 设置坐标轴属性
    ax.set_xlabel("Time (s, relative to peak)", fontsize=6)
    ax.set_ylabel("Rear SEAT_CENTER (dB-A)", fontsize=6)
    ax.set_title("Rear Seat Center Noise Graph", fontsize=6)
    ax.tick_params(axis='both', labelsize=5)
    ax.grid(True)
    ax.legend(fontsize=5)
    ax.xaxis.set_major_locator(MaxNLocator(integer=True))

    # 横坐标范围控制：保留 peak 前数据 + 限制右侧至 terminate_sec
    left_limit = time_list[start_idx] - t0
    right_limit = float(terminate_sec) if terminate_sec is not None else shifted_time[-1]
    ax.set_xlim(left=left_limit, right=right_limit)

def conclusion_writer(csv_path, peak, peak_after_one_min, before_one_min_val, after_one_min_val):
    csv_path = Path(csv_path)
    folder= csv_path.parent
    xlsx_path = folder / "result.xlsx"

    wb = load_workbook(xlsx_path)
    ws = wb["conclusion"]

    stem = csv_path.stem
    parts = stem.split("_")

    if 'FL' in parts:
        fuel_key = 'FL'
    elif {'7', '8'}.issubset(parts):
        fuel_key = '7/8'
    elif {'3', '4'}.issubset(parts):
        fuel_key = '3/4'
    else:
        raise ValueError(f"fuel type recognition error: {stem}")

    valid_acc_vols = {'02G', '03G'}
    acc_vol_candidates = valid_acc_vols.intersection(parts)
    if acc_vol_candidates:
        acc_vol = acc_vol_candidates.pop()
    else:
        raise ValueError(f"fuel volume recognition error: {stem}")

    judge = "G" if (peak < before_one_min_val and peak_after_one_min < after_one_min_val) else "R"

    acc_to_row = {
        '02G': 2,
        '03G': 5
    }

    fuel_to_offset = {
        'FL': 0,     # FULL
        '7/8': 1,     # 7/8
        '3/4': 2      # 3/4
    }

    base_row = acc_to_row.get(acc_vol.upper())
    offset = fuel_to_offset.get(fuel_key.upper())

    if base_row is None or offset is None:
        raise ValueError(f"error: acc={acc_vol}, fuel={fuel_key}")

    writing_row = base_row + offset

    ws.cell(row=writing_row, column=3, value=round(peak,1))
    ws.cell(row=writing_row, column=4, value=before_one_min_val)
    ws.cell(row=writing_row, column=5, value=round(peak_after_one_min,1))
    ws.cell(row=writing_row, column=6, value=after_one_min_val)
    ws.cell(row=writing_row, column=7, value=judge)

    wb.save(xlsx_path)

def generate_graph_from_csv(csv_path, before_one_val, after_one_val, terminate_sec, save_name=None, save=False):
    df = pd.read_csv(csv_path)

    expected_columns = ["Time(s)", "ACC_X(G)", "REAR SEAT_CENTER(dB-A)"]
    for col in expected_columns:
        if col not in df.columns:
            raise ValueError(f"Missing column: {col}, please check the .csv file")

    time_list = df["Time(s)"].tolist()
    acc_list = df["ACC_X(G)"].tolist()
    noise_list = df["REAR SEAT_CENTER(dB-A)"].tolist()

    delta_acc_list = [acc_list[i] - acc_list[i - 1] for i in range(1, len(acc_list))]
    start_point = delta_acc_list.index(max(delta_acc_list)) + 1 
    peak_point = acc_list.index(max(acc_list[start_point:]))

    print(f"peak point index = {peak_point}, Time = {time_list[peak_point]}")

    fig, axs = plt.subplots(2, 1, figsize=(5, 4))
    fig.suptitle(save_name or "Unnamed", fontsize=7)
    plot_acceleration_graph(axs[0], time_list, acc_list, peak_point)
    plot_rear_center_noise(
        axs[1],
        time_list,
        noise_list,
        peak_point,
        before_one_min_val=before_one_val,
        after_one_min_val=after_one_val,
        terminate_sec=terminate_sec
    )

    # 计算终止索引
    after_one_min_point = peak_point + 25
    if after_one_min_point >= len(noise_list):
        after_one_min_point = len(noise_list) - 1  # 防越界

    if terminate_sec is None:
        end_idx = len(time_list)
    else:
        terminate = time_list[peak_point] + float(terminate_sec)
        end_idx = bisect.bisect_left(time_list, terminate)
        end_idx = min(end_idx, len(time_list))  # 防越界

    conclusion_writer(
        csv_path,
        peak=max(noise_list[peak_point:min(peak_point + 24, len(noise_list))]),
        peak_after_one_min=max(noise_list[after_one_min_point:end_idx]),
        before_one_min_val=before_one_val,
        after_one_min_val=after_one_val 
    )

    plt.tight_layout()

    if save:
        folder = os.path.dirname(csv_path)
        save_folder = os.path.join(folder, "figures")
        os.makedirs(save_folder, exist_ok=True)
        fig_name = f'{save_name or "figure"}.png'
        save_path = os.path.join(save_folder, fig_name)
        fig.savefig(save_path, dpi=300)

    plt.close(fig)

def xlsx_init(path):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "conclusion"
    ws2 = wb.create_sheet(title="result")

    headers1 = ["加速度", "容量", "車内音 ~1s (dB-A)", "目標値(dB-A)", "車内音 1s~ (dB-A)", "目標値(dB-A)", "判定"]
    headers2 = ["加速度", "", "Full", "",  "7/8", "",  "3/4"]

    ws1.append(headers1)
    ws2.append(headers2)

    ws1['A2'] = "0.2G"
    ws1['A5'] = "0.3G"
    ws1.merge_cells('A2:A4')
    ws1.merge_cells('A5:A7')

    ws1['B2'] = "FULL"
    ws1['B3'] = "7/8"
    ws1['B4'] = "3/4"
    ws1['B5'] = "FULL"
    ws1['B6'] = "7/8"
    ws1['B7'] = "3/4"

    for col in ['C', 'E']:
        ws1.column_dimensions[col].width = 17
    for col in ['D', 'F']:
        ws1.column_dimensions[col].width = 13

    thin = Side(border_style="thin", color="000000")
    for row in ws1.iter_rows(min_row=1, max_row=7, min_col=1, max_col=7):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws2.row_dimensions[2].height = 5
    ws2.row_dimensions[4].height = 5
    ws2.row_dimensions[3].height = 230
    ws2.row_dimensions[5].height = 230

    for col in ['C', 'E', 'G']:
        ws2.column_dimensions[col].width = 51

    for col in ['B', 'D', 'F']:
        ws2.column_dimensions[col].width = 0.5

    ws2['A3'] = "0.2G"
    ws2['A5'] = "0.3G"

    center_alignment = Alignment(horizontal='center', vertical='center')

    ws1['A2'].alignment = center_alignment
    ws1['A5'].alignment = center_alignment

    ws2['A3'].alignment = center_alignment
    ws2['A5'].alignment = center_alignment
    ws2['C1'].alignment = center_alignment
    ws2['E1'].alignment = center_alignment
    ws2['G1'].alignment = center_alignment

    for row in ws2.iter_rows(min_row=1, max_row=5, min_col=1, max_col=7):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    save_path = f'{path}/result.xlsx'
    wb.save(save_path)

def fig_writer(path):
    path = Path(path)

    xlsx_path = path / "result.xlsx"
    img_folder_path = path / "figures"

    if not xlsx_path.exists():
        raise FileNotFoundError(f"could not find Excel file:{xlsx_path}")
    if not img_folder_path.exists():
        raise FileNotFoundError(f"could not load the image:{img_folder_path}")

    wb = load_workbook(xlsx_path)
    ws = wb["result"]

    acc_to_row = {
        '02G': 3,
        '03G': 5
    }
    fuel_to_col = {
        'FL': 'C',
        '7/8': 'E',
        '3/4': 'G'
    }

    for img_path in img_folder_path.iterdir():
        if not img_path.is_file() or img_path.suffix.lower() != ".png":
            continue

        img = Image(img_path)
        img.width = 370
        img.height = 300

        stem = img_path.stem
        parts = stem.split("_")

        if 'FL' in parts:
            fuel_key = 'FL'
        elif {'7', '8'}.issubset(parts):
            fuel_key = '7/8'
        elif {'3', '4'}.issubset(parts):
            fuel_key = '3/4'
        else:
            raise ValueError(f"fuel type recognition error: {stem}")

        valid_acc_vols = {'02G', '03G'}
        acc_vol_candidates = valid_acc_vols.intersection(parts)
        if acc_vol_candidates:
            acc_vol = acc_vol_candidates.pop()
        else:
            raise ValueError(f"fuel volume recognition error: {stem}")

        import_row = acc_to_row.get(acc_vol.upper())
        import_col = fuel_to_col.get(fuel_key.upper())

        if import_row is None or import_col is None:
            raise ValueError(f"could not find import position, acc={acc_vol}, fuel={fuel_key}, 文件名={stem}")

        import_position = f"{import_col}{import_row}"

        ws.add_image(img, import_position)

    wb.save(xlsx_path)

def processer(path, before_one_val, after_one_val, end_time, progress_callback):
    folder_path = Path(path)
    xlsx_init(path)

    csv_files = [f for f in folder_path.iterdir() if f.is_file() and f.suffix.lower() == ".csv"]
    total = len(csv_files)

    if total == 0:
        raise FileNotFoundError("could not find csv file")

    for i, file_path in enumerate(csv_files, 1):
        try:
            stem = file_path.stem
            generate_graph_from_csv(file_path, before_one_val, after_one_val, end_time, save_name=stem, save=True)
        except Exception as e:
            print(f"csv process failure: {file_path.name}:{e}")

        progress = int(i / total * 90)  
        progress_callback.emit(progress)

    # 图像写入 Excel
    try:
        fig_writer(path)
        progress_callback.emit(100)
    except Exception as e:
        print(f"figure import failure: {e}")
        progress_callback.emit(100)

class WorkerThread(QThread):
    progress = pyqtSignal(int)

    def __init__(self, path, before_one_val, after_one_val, end_time):
        super().__init__()
        self.path = path
        self.before_one_val = before_one_val
        self.after_one_val = after_one_val
        self.end_time = end_time

    def run(self):
        processer(
            path=self.path,
            before_one_val=self.before_one_val,
            after_one_val=self.after_one_val,
            end_time=self.end_time,
            progress_callback=self.progress
        )

# main gui
class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("揺動音自動処理ツール")
        self.setAcceptDrops(True)
        self.setFixedSize(420, 300)

        self.folder_path = None

        self.folder_label = QLabel(".CSVを導入してください")
        self.folder_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.folder_label.setStyleSheet("border: 2px dashed gray; padding: 40px;")

        # 基準値入力1
        self.ref_value_label = QLabel("開始から1sまでの基準値:")
        self.ref_value_input = QLineEdit()
        self.ref_value_input.setText("54")
        self.ref_value_input.setFixedWidth(60)

        # 基準値输入2
        self.ref_value_label2 = QLabel("1s以後の基準値:")
        self.ref_value_input2 = QLineEdit()
        self.ref_value_input2.setText("34")
        self.ref_value_input2.setFixedWidth(60)

        # 終了時間
        self.end_time_label = QLabel("終了時間を設定してください(単位-秒):")
        self.end_time_select = QComboBox()
        self.end_time_select.addItems(["None", "5", "4", "3", "2", "1"])
        self.end_time_select.setCurrentIndex(0)
        self.end_time_select.setFixedWidth(80)

        # ボタン
        self.start_button = QPushButton("処理開始")
        self.start_button.clicked.connect(self.start_processing)

        self.quit_button = QPushButton("終了")
        self.quit_button.clicked.connect(self.close)

        # プログレスバー
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)

        # レイアウト設定
        layout = QVBoxLayout()
        layout.addWidget(self.folder_label)

        # 输入框布局
        ref_layout = QHBoxLayout()
        ref_layout.addWidget(self.ref_value_label)
        ref_layout.addWidget(self.ref_value_input)

        ref_layout2 = QHBoxLayout()
        ref_layout2.addWidget(self.ref_value_label2)
        ref_layout2.addWidget(self.ref_value_input2)

        time_layout = QHBoxLayout()
        time_layout.addWidget(self.end_time_label)
        time_layout.addWidget(self.end_time_select)

        layout.addLayout(ref_layout)
        layout.addLayout(ref_layout2)
        layout.addLayout(time_layout)

        layout.addWidget(self.progress_bar)

        # ボタン配置
        buttons = QHBoxLayout()
        buttons.addWidget(self.start_button)
        buttons.addWidget(self.quit_button)

        layout.addLayout(buttons)
        self.setLayout(layout)

    # 拖拽进入事件
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    # 拖拽释放事件
    def dropEvent(self, event):
        urls = event.mimeData().urls()
        if urls:
            folder = Path(urls[0].toLocalFile())
            if folder.is_dir():
                self.folder_path = folder
                self.folder_label.setText(f"✅ フォルダ読み込み完了")
            else:
                self.folder_label.setText("⚠️ フォルダをインポートしてください")

    # 启动主线程处理
    def start_processing(self):
        if not self.folder_path:
            self.folder_label.setText("⚠️ フォルダ無効")
            return

        before_one_val_text = self.ref_value_input.text()
        after_one_val_text = self.ref_value_input2.text()

        try:
            self.before_one_val = float(before_one_val_text)
            self.after_one_val = float(after_one_val_text)
        except ValueError:
            self.folder_label.setText("⚠️ 基準値は数値で入力してください")
            return

        end_time = self.end_time_select.currentText()
        if end_time == 'None':
            self.end_time = None
        else:
            self.end_time = int(end_time)


        self.progress_bar.setValue(0)
        self.worker = WorkerThread(self.folder_path, 
                                   self.before_one_val, 
                                   self.after_one_val, 
                                   self.end_time)
        self.worker.progress.connect(self.progress_bar.setValue)
        self.worker.finished.connect(lambda: self.folder_label.setText("✅ 処理完了"))
        self.folder_label.setText("⏳ 処理中")
        self.worker.start()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.processEvents()
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
